""" Copyright (c) 2022 Cisco and/or its affiliates.
This software is licensed to you under the terms of the Cisco Sample
Code License, Version 1.1 (the "License"). You may obtain a copy of the
License at
           https://developer.cisco.com/docs/licenses
All use of the material herein must be in accordance with the terms of
the License. All rights not expressly granted by the License are
reserved. Unless required by applicable law or agreed to separately in
writing, software distributed under the License is distributed on an "AS
IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
or implied.
"""

#Imports
import os
import requests
import json
from dotenv import load_dotenv
import openpyxl

#Environment Variables
load_dotenv()
MERAKI_BASE_URL = os.environ['MERAKI_BASE_URL']

#Headers
headers = {
    "Content-Type": "application/json",
    "Accept": "application/json",
    "X-Cisco-Meraki-API-Key": os.environ["MERAKI_API_KEY"]
}


def read_input_file():
    dataframe = openpyxl.load_workbook("New-Input-Format.xlsx" )

    excel_sheets=[]
    global templates_list
    templates_list={}
    for sheet in dataframe.worksheets:
        excel_sheets.append(sheet.title)
        templates_list[sheet.title]=[]
    
    for sheet in excel_sheets:
        dataframe1 = dataframe[sheet]
        templates_list[sheet]={}
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            if str(col[1].value) != "None":
                templates_list[sheet][str(col[0].value)]=str(col[1].value)
    
    # print(json.dumps(templates_list, indent=2))



def get_saml_and_security(orgId):
    try:
        saml_url=f"/organizations/{orgId}/saml"
        saml_response = requests.get(MERAKI_BASE_URL+saml_url, headers=headers).json()

        login_security_url=f"/organizations/{orgId}/loginSecurity"
        login_security_response = requests.get(MERAKI_BASE_URL+login_security_url, headers=headers).json()
        login_security_response['enabled']=saml_response['enabled']

        compliant_list=[]
        non_compliant_list=[]

        # print(json.dumps(login_security_response, indent=2))

        if templates_list['ORG-SEC']['SSO'] != login_security_response['enabled']:
            non_compliant_list.append({'SSO Enabled:':login_security_response['enabled']})
        else:
            compliant_list.append({'SSO Enabled:':login_security_response['enabled']})
        
        if int(templates_list['ORG-SEC']['Password expiration']) != login_security_response['passwordExpirationDays']:
            non_compliant_list.append({'Password Expiration (Days):':login_security_response['passwordExpirationDays']})
        else:
            compliant_list.append({'Password Expiration (Days):':login_security_response['passwordExpirationDays']})
        
        if int(templates_list['ORG-SEC']['Used password']) != login_security_response['numDifferentPasswords']:
            non_compliant_list.append({'Used Password:':login_security_response['numDifferentPasswords']})
        else:
            compliant_list.append({'Used Password:':login_security_response['numDifferentPasswords']})
        
        if templates_list['ORG-SEC']['Strong password'] != login_security_response['enforceStrongPasswords']:
            non_compliant_list.append({'Strong Password:':login_security_response['enforceStrongPasswords']})
        else:
            compliant_list.append({'Strong Password:':login_security_response['enforceStrongPasswords']})
        
        if int(templates_list['ORG-SEC']['Account Lockout']) != login_security_response['enforceAccountLockout']:
            non_compliant_list.append({'Account Lockout:':login_security_response['enforceAccountLockout']})
        else:
            compliant_list.append({'Account Lockout:':login_security_response['enforceAccountLockout']})
        
        if int(templates_list['ORG-SEC']['Idle Timeout']) != login_security_response['idleTimeoutMinutes']:
            non_compliant_list.append({'Idle Timeout (Minutes):':login_security_response['idleTimeoutMinutes']})
        else:
            compliant_list.append({'Idle Timeout (Minutes):':login_security_response['idleTimeoutMinutes']})
        
        if templates_list['ORG-SEC']['Two-factor authentication'] != login_security_response['enforceTwoFactorAuth']:
            non_compliant_list.append({'Two-factor Authentication:':login_security_response['enforceTwoFactorAuth']})
        else:
            compliant_list.append({'Two-factor Authentication:':login_security_response['enforceTwoFactorAuth']})

        
        return compliant_list,non_compliant_list
    except Exception as e:
            print("Exception in get_saml: " + str(e))

def get_networks_ssids(org_id):
    try:
        url = MERAKI_BASE_URL+ f"/organizations/{org_id}/networks"
        networks = requests.get(url, headers=headers).json()
        non_compliant_list=[]
        
        for network in networks:
                url=f"/networks/{network['id']}/wireless/ssids"
                ssid_response=requests.get(MERAKI_BASE_URL+url, headers=headers).json()
            
                if int(len(ssid_response)) == int(templates_list['SSIDs']['Number of SSID']):
                    n_name='HXXXX-hotel x'
                    n_name=n_name.split('-')
                    n_name=n_name[1]
                    ssid_names=[]
                    for i in ssid_response:
                        ssid_names.append(i['name'])

                    if n_name and f'{n_name} Meeting' and f'{n_name} Staff' in ssid_names:
                        #Now check the configs for each 
                        notes=''
                        for i in ssid_response:
                            if i['name'] == f'{n_name} Staff':
                                if ('radiusServers' in i and len(i['radiusServers']) == 2) and ('radiusAccountingServers' in i):
                                    if templates_list['Hotel Staff SSID']['Radius primary server'] ==i['radiusServers'][0]['host'] and templates_list['Hotel Staff SSID']['Radius secondary server'] ==i['radiusServers'][1]['host'] and int(templates_list['Hotel Staff SSID']['Authentication port']) ==i['radiusServers'][1]['port'] and int(templates_list['Hotel Staff SSID']['Authentication port']) ==i['radiusServers'][0]['port'] and int(templates_list['Hotel Staff SSID']['Accounting port']) ==i['radiusAccountingServers'][0]['port']:
                                    
                                        if templates_list['Hotel Staff SSID']['visible']== i['visible'] and templates_list['Hotel Staff SSID']['encryptionMode']== i['encryptionMode']:
                                            print('Compliant')
                                        else:
                                            notes=notes+f'Non-compliant configuration for SSID {i["name"]}, '
                                    else:
                                        notes=notes+f'Non-compliant configuration for SSID {i["name"]}, '
                                else:
                                    notes=notes+f'Non-compliant configuration for SSID {i["name"]}, '
                            elif i['name'] == f'{n_name} Meeting':
                                print("in meeting")
                                if str(templates_list['Meeting Area SSID']['visible'])== str(i['visible']) and  templates_list['Meeting Area SSID']['splashPage']== i['splashPage']:
                                    print('Compliant')
                                else:
                                    notes=notes+f'Non-compliant configuration for SSID {i["name"]}, '
                            elif i['name'] == n_name:
                                if str(templates_list['Guest SSID']['visible'])== str(i['visible']) and  templates_list['Guest SSID']['splashPage']== i['splashPage']:
                                    print('Compliant')
                                else:
                                    notes=notes+f'Non-compliant configuration for SSID {i["name"]}'
                        if notes!='':
                            non_compliant_list.append({f'{n_name} ({len(ssid_response)})':notes})
                    else:
                            non_compliant_list.append({f'{network["name"]} ({len(ssid_response)})':'SSIDs naming conventions not followed'})

                else:
                    non_compliant_list.append({f'{network["name"]} ({len(ssid_response)})':'correct number of SSIDs not followed'})

        return non_compliant_list
    except Exception as e:
            print("Exception in get_networks_ssids: " + str(e))


def get_networks_vlans(org_id):
    try:
        url = MERAKI_BASE_URL+ f"/organizations/{org_id}/networks"
        networks = requests.get(url, headers=headers).json()
        non_compliant_list=[]
        vlans_number = int(templates_list['VLANs']['Number of VLANs'])
        for network in networks:
                url=f"/networks/{network['id']}/vlanProfiles"
                vlan_response=requests.get(MERAKI_BASE_URL+url, headers=headers).json()
                if  "errors" in vlan_response:
                    non_compliant_list.append({network['name']:"NA"})
                elif len(vlan_response) != 1:
                        for profile in vlan_response:
                            non_compliant_list.append({network['name']+f" ({(str(len(vlan_response)))})":profile['name']+f" ({str(len(profile['vlanNames']))})"})

                else:
                        notes=''
                    
                        k=[]
                        for tt in templates_list['VLANs'].keys():
                            if tt != "Number of VLANs":
                                k.append(tt)

                        if len(vlan_response[0]['vlanNames']) == vlans_number:
                            for m in vlan_response[0]['vlanNames']:
                                    if m['name'] in k and templates_list['VLANs'][m['name']] == m['vlanId']:
                                        print(f'VLAN {m["name"]} compliant')
                                    else:
                                        print(f'VLAN {m["name"]} NOT compliant')
                                        notes=notes+f'VLAN {m["name"]} NOT compliant, '
                            non_compliant_list.append({network['name']+f" ({(str(len(vlan_response)))})":profile['name']+f" ({str(len(profile['vlanNames']))}) ({notes})"})
                                        
                        else: 
                            non_compliant_list.append({network['name']+f" ({(str(len(vlan_response)))})":profile['name']+f" ({str(len(profile['vlanNames']))})"})

        return non_compliant_list
    except Exception as e:
            print("Exception in get_netwrok_vlans: " + str(e))

#Get all devices in an organization
def get_org_devices(org_id):
    try:
        url= MERAKI_BASE_URL + f"/organizations/{org_id}/networks"
        networks = requests.get(url,headers=headers).json()
        net_ids_names=[]
        for network in networks:
            entry={'name':network['name'],
                    'id':network['id']
                  }
            net_ids_names.append(entry)

        sw_non_compliant_list=[]
        ap_non_compliant_list=[]
        net_non_compliant_list=[]

        url= MERAKI_BASE_URL + f"/organizations/{org_id}/devices?productTypes[]=switch"
        devices_sw = requests.get(url,headers=headers).json()
        # print(json.dumps(devices_sw, indent=2))
        
        url= MERAKI_BASE_URL + f"/organizations/{org_id}/devices?productTypes[]=wireless"
        devices_ap = requests.get(url,headers=headers).json()
        # print(json.dumps(devices_wi, indent=2))

        for net in net_ids_names:
            n=net['name'].split('-')
            if len(n) !=2:
                # print('non-compliant')
                net_non_compliant_list.append(net['name'])
            else:
                if len(n[0])== 5 and n[0][0]=='H':
                    print("compliant")
                else:
                    # print('non-compliant')
                    net_non_compliant_list.append(net['name'])

        for n in net_ids_names:
            
                for sw in devices_sw:
                    if sw['networkId'] == n['id']:
                        m=sw['name'].split('-')
                        if len(m) !=4:
                            print('non-compliant')
                            sw_non_compliant_list.append({sw['name']:n['name']})
                        else:
                            if len(m[0])==2 and len(m[1])==5 and m[1][0]=='H' and m[2]=='SW':
                                print("compliant")
                            else:
                                print('non-compliant')
                                sw_non_compliant_list.append({sw['name']:n['name']})

                for ap in devices_ap:
                    if ap['networkId'] == n['id']:
                        m=ap['name'].split('-')
                        if len(m) !=4:
                            print('non-compliant')
                            ap_non_compliant_list.append({sw['name']:n['name']})
                        else:
                            if len(m[0])==2 and len(m[1])==5 and m[1][0]=='H' and m[2]=='AP':
                                print("compliant")
                            else:
                                print('non-compliant')
                                ap_non_compliant_list.append({sw['name']:n['name']})
                
        return sw_non_compliant_list,ap_non_compliant_list,net_non_compliant_list
    except Exception as e:
            print("Exception in get_org_devices: " + str(e))

